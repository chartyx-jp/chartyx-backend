import matplotlib.pyplot as plt
plt.rcParams['font.family'] = 'MS Gothic' 

import xgboost as xgb
import shap
import pandas as pd
import numpy as np
from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from apps.common.utils import Utils
from pathlib import Path
from operator import itemgetter
from sklearn.metrics import mean_absolute_error, root_mean_squared_error, r2_score
from apps.common.app_initializer import DjangoAppInitializer
from apps.ai.ai_models.base_booster import BaseBoosterModel
from django.conf import settings

def auto_save_plot(model_name: str = None):
    def decorator(func):
        def wrapper(self, *args, **kwargs):
            result = func(self, *args, **kwargs)
            self.log.info("decorator: auto_save_plot")
            plot_name = getattr(self, "_plot_name", None)
            if plot_name:
                self.save_plot(plot_name=plot_name, model_name=model_name)
                plt.show()
                plt.close()  # OOM対策
            else:
                self.log.warning("plot_name が指定されていないため、保存をスキップしました")
            return result
        return wrapper
    return decorator

class BoosterEvaluatorV2(DjangoAppInitializer):
    """
    XGBoost Booster形式のモデル評価クラス。
    - BaseBoosterModel を受け取り、特徴量重要度分析や予測評価を行う。
    """

    def __init__(self, model: BaseBoosterModel, group_label: str = "", ticker_name: str = "", *args, **kwargs) -> None:
        super().__init__(*args, **kwargs)
        self.__model = model
        self.__booster = self.__model.booster
        self.group_label = group_label
        self.ticker_name = ticker_name
        self._plot_name = None

    def print_feature_importance(self, top_n: int = 20):
        score = self.__booster.get_score(importance_type='gain')
        sorted_items = sorted(score.items(), key=lambda x: x[1], reverse=True)
        self.log.info(f"[特徴量重要度TOP{top_n}（gain）] for {self.group_label}/{self.ticker_name}")
        for feature, importance in sorted_items[:top_n]:
            self.log.info(f"{feature}: {importance:.4f}")

    @auto_save_plot()
    def plot_feature_importance(self, top_n: int = 20):
        score = self.__booster.get_score(importance_type='gain')
        sorted_items = sorted(score.items(), key=lambda x: x[1], reverse=True)[:top_n]
        self._plot_name = f"Feature Importance (gain)\n{self.__model.model_name}"

        features = [f for f, _ in sorted_items]
        importances = [s for _, s in sorted_items]

        # 動的にグラフサイズ調整
        height = max(6, 0.45 * len(features))
        plt.figure(figsize=(10, height))
        bars = plt.barh(features[::-1], importances[::-1], color=plt.cm.Blues(np.linspace(0.3, 0.7, len(features))))
        plt.title(self._plot_name, fontsize=15)
        plt.xlabel("Gain", fontsize=13)
        plt.yticks(fontsize=11)

        # 値をバーの右に表示
        for bar, val in zip(bars, importances[::-1]):
            plt.text(bar.get_width(), bar.get_y() + bar.get_height()/2, f"{val:.3f}", va='center', ha='left', fontsize=9)
        plt.tight_layout()
        plt.grid(axis='x', linestyle=':', alpha=0.4)

    @auto_save_plot()
    def plot_prediction_vs_actual(self, y_true: pd.Series, y_pred: pd.Series):
        self._plot_name = f"Predicted vs Actual\n{self.group_label}/{self.ticker_name}"
        corr = np.corrcoef(y_true, y_pred)[0, 1]

        plt.figure(figsize=(8, 8))
        plt.scatter(y_true, y_pred, alpha=0.4, label="Data Points")
        min_val = min(y_true.min(), y_pred.min())
        max_val = max(y_true.max(), y_pred.max())
        margin = (max_val - min_val) * 0.05
        plt.plot([min_val, max_val], [min_val, max_val], 'r--', lw=1, label="y=x (理想)")

        plt.xlim(min_val - margin, max_val + margin)
        plt.ylim(min_val - margin, max_val + margin)
        plt.xlabel("Actual", fontsize=12)
        plt.ylabel("Predicted", fontsize=12)
        plt.title(f"{self._plot_name}\n相関係数: {corr:.3f}", fontsize=15)
        plt.grid(True, linestyle=':', alpha=0.5)
        plt.legend(loc='upper right', fontsize=10)
        # 統計情報を図上に表示
        mae = mean_absolute_error(y_true, y_pred)
        rmse = root_mean_squared_error(y_true, y_pred)
        plt.annotate(f"MAE={mae:.4f}\nRMSE={rmse:.4f}", xy=(0.02, 0.98), xycoords='axes fraction', 
                    fontsize=11, ha='left', va='top',
                    bbox=dict(boxstyle="round,pad=0.3", fc="w", ec="k", lw=0.7, alpha=0.6))
        plt.tight_layout()

    @auto_save_plot()
    def plot_error_histogram(self, y_true: pd.Series, y_pred: pd.Series):
        errors = y_pred - y_true
        self._plot_name = f"Prediction Error Histogram\n{self.group_label}/{self.ticker_name}"

        plt.figure(figsize=(12,6))
        bins = min(max(10, len(errors) // 5), 50)
        n, bins, patches = plt.hist(errors, bins=bins, color='skyblue', edgecolor='black', alpha=0.85)
        plt.axvline(0, color='red', linestyle='--', linewidth=1, label="Zero Error")
        mean_error = errors.mean()
        plt.axvline(mean_error, color='orange', linestyle='--', linewidth=1, label="Mean Error")
        plt.text(mean_error, plt.ylim()[1]*0.9, f"mean={mean_error:.4f}", color='orange')
        # 統計値
        plt.title(self._plot_name + f"\n歪度={errors.skew():.2f}, 尖度={errors.kurtosis():.2f}", fontsize=14)
        plt.xlabel("Prediction Error", fontsize=12)
        plt.ylabel("Frequency", fontsize=12)
        plt.xticks(fontsize=11)
        plt.grid(True, axis='x', linestyle=':', alpha=0.4)
        plt.legend()
        plt.tight_layout()

    @auto_save_plot()
    def plot_shap_summary(self, X: pd.DataFrame):
        try:
            self._plot_name = f"SHAP Summary\n{self.group_label}/{self.ticker_name}"
            explainer = shap.TreeExplainer(self.__booster)
            shap_values = explainer.shap_values(X)
            plt.figure(figsize=(14, max(6, 0.4 * len(X.columns))))
            shap.summary_plot(shap_values, X, show=False)
            plt.title(self._plot_name, fontsize=15)
            plt.tight_layout()
        except Exception as e:
            self.log.error(f"SHAPプロット中にエラーが発生しました: {e}")

    def evaluate_predictions(self, X: pd.DataFrame, y_true: pd.Series):
        try:
            dmatrix = xgb.DMatrix(X)
            y_pred = self.__booster.predict(dmatrix)
            self.plot_prediction_vs_actual(y_true, y_pred)
            self.plot_error_histogram(y_true, y_pred)
            mae = mean_absolute_error(y_true, y_pred)
            rmse = root_mean_squared_error(y_true, y_pred)
            mape = self.safe_mape(y_true, y_pred)
            r2 = r2_score(y_true, y_pred)
            return {
                "mae": mae,
                "rmse": rmse,
                "r2": r2,
                "mape": mape}
        except Exception as e:
            self.log.error(f"予測評価中にエラーが発生しました: {e}")
            return None

    def safe_mape(self, y_true, y_pred):
        mask = y_true != 0
        return (abs((y_true[mask] - y_pred[mask]) / y_true[mask])).mean() * 100

    def evaluate(self, X: pd.DataFrame, y: pd.Series):
        if X.empty or y.empty:
            self.log.warning("評価対象データが空のため、スキップします。")
            return
        self.plot_feature_importance()
        self.plot_shap_summary(X)
        results = self.evaluate_predictions(X, y)
        if results:
            mae, rmse, mape, r2 = itemgetter("mae", "rmse", "mape", "r2")(results)
            self.log.info(f"[評価結果] {self.group_label}/{self.ticker_name} → MAE: {mae:.2f}, RMSE: {rmse:.2f}, MAPE: {mape:.2f}%, R²: {r2:.4f}")
            return results

    def save_to_excel(self, model_name: str, metrics: dict, path: Path = settings.LOG_DIR / "evaluation_log.xlsx"):

        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        row = {
            "timestamp": now,
            "model": model_name,
            "group": self.group_label,
            "ticker": self.ticker_name,
            **metrics
        }
        new_df = pd.DataFrame([row])
        if path.exists():
            existing = pd.read_excel(path)
            df = pd.concat([existing, new_df], ignore_index=True)
        else:
            df = new_df
        df.to_excel(path, index=False)

        # --- ★ここからテーブル化＋列幅最適化 ---
        wb = load_workbook(path)
        ws = wb.active

        # テーブル範囲（A1から最終セルまで）
        table_ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        table_name = "ResultTable"
        # すでにテーブルがあれば削除（上書き用）
        if table_name in ws.tables:
            del ws.tables[table_name]
        tab = Table(displayName=table_name, ref=table_ref)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                            showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)

        # 列幅を自動調整
        for col in ws.columns:
            max_length = 0
            col_name = col[0].column_letter
            for cell in col:
                try:
                    value = str(cell.value)
                    if value:
                        max_length = max(max_length, len(value))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[col_name].width = adjusted_width

        wb.save(path)
        
    def save_plot(self, plot_name: str, model_name: str = None):
        """
        モデル名とプロット名をもとに、グラフ画像をファイル保存する。
        保存先: settings.RESEARCH_DATA_DIR / model_name / safe_plot.png
        """
        model_dir = model_name if model_name else self.__model.model_name
        dir_path = Path(settings.RESEARCH_DATA_DIR) / model_dir
        dir_path.mkdir(parents=True, exist_ok=True)
        safe_plot = Utils.safe_filename_component(plot_name)
        full_path = dir_path / safe_plot
        plt.savefig(full_path, dpi=300, bbox_inches='tight')
        self.log.info(f"[保存完了] プロット保存: {full_path}")

